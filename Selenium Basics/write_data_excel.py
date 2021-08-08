import openpyxl

file_source = "Login_Data.xlsx"
open_file = openpyxl.load_workbook(file_source)
working_sheet = open_file["Write"]

for r in range(2, 7):
    for c in range(1, 4):
        working_sheet.cell(row=2, column=1).value = "This is Pranto"
        working_sheet.cell(row=3, column=1).value = "This is Maksud"
        working_sheet.cell(row=6, column=3).value = "This is Yra"

open_file.save(file_source)
