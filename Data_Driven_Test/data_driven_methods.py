import openpyxl


def get_row_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.max_row


def get_column_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.max_column


def read_data(file, sheet, reading_rows, reading_columns):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.cell(row=reading_rows, column=reading_columns).value


def writing_data(file, sheet, writing_rows, writing_columns, data):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    worksheet.cell(row=writing_rows, column=writing_columns).value = data
    workbook.save(file)
